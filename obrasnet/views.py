from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_protect
from django.contrib.staticfiles import finders
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail

# from obrasnet.forms import RegisterForm

from openpyxl import load_workbook

from obrasnet.models import *
from project import settings

# Home page for url '/'

def index(request):
    return render(request, 'index.html')

# Login, Logout, Register page

## Register
def register(request, *args, **kwargs):
    if request.method == 'POST':
        username = request.POST['username']
        first_name = request.POST['fname']
        last_name = request.POST['lname']
        email = request.POST['email']
        pass1 = request.POST['pass1']
        pass2 = request.POST['pass2']

        if User.objects.filter(username=username):
            messages.error(request, 'Username already exist! Please try some other name.')
        
        if User.objects.filter(email=email):
            messages.error(request, 'Email already registered!')
        
        if len(username) > 20:
            messages.error(request, "Username must be under 20 characters")

        if pass1 != pass2:
            messages.error(request, "Passwords do not match")

        if not username.isalnum():
            messages.error(request, "Username must be Alpha-numeric")

        myuser = User.objects.create_user(username, email, pass1)
        myuser.first_name = first_name
        myuser.last_name = last_name

        myuser.save()

        messages.success(request, "Your account has been successfully created. We have sent you a confirmation email, please confirm your email in order to activate your account.")

        # Welcome Email
        # subject = "Welcome to Obrasnet!"
        # message = "Hello " + myuser.first_name + "!! \n" + "Thank you for visiting our website. \n We have to also sent you a confirmation email, please confirm your email address in order to activate your account. \n\n Thank you. \n Flaviano Souza."

        # from_email = settings.EMAIL_HOST_USER
        # to_list = [myuser.email]

        # send_mail(subject, message, from_email, to_list, fail_silently=True)
        
        return redirect('signin')

    return render(request, 'auth/register.html')

## Login
def signin(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return render(request, 'index.html', {'username': user.username})
        else:
            messages.error(request, 'Invalid username or password')
        
    return render(request, 'auth/login.html')

## Logout
def signout(request):
    logout(request)

    # messages.success(request, "Log out successfully")

    return redirect('home')

# Create a new Requirement for url '/requirements'
## In this, client makes a reqruirement for the design.

def requirements(request):
    return render(request, "requirements.html")

# Recommend the designs based on the requirements.
## 

@csrf_protect
def design_recommendations(request):
    if request.method == 'POST':
        data = {
            'title': request.POST.get('title'),
            'content': request.POST.get('content'),
        }

        return render(request, "design_recommendations.html", data)


# Update Scrapping Material Datas in material table
## In Excel, last column is state of material. 0: deleted, 1: existed, 2: created
## In Table, If state is True, material is used. If False, deleted.

def scrap_history(request):
    categories = Category.objects.all()


    for category in categories:
        filename = finders.find('upload/scrap_data/' + category.name + '.xlsx')

        wb = load_workbook(filename)

        sheets = category.sheetname.split(',')
        
        for sh in sheets:
            sheet = wb[sh]
            i = 2
            for row in sheet.iter_rows(min_row=2, min_col=2):
                if row[10].value == 2:
                    material = Material(
                        title = row[0].value,
                        image = row[1].value,
                        specification = row[2].value,
                        ref = row[3].value,
                        price = row[4].value,
                        description = row[5].value,
                        advantage = row[6].value,
                        documentation = row[7].value,
                        features = row[8].value,
                        url = row[9].value,
                        category_id = category.id,
                        state = 1
                    )

                    material.save()

                    sheet.cell(row=i, column=12).value = 1
                
                i = i + 1
            
            wb.save(filename)
            
    return render(request, 'index.html')
