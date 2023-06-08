from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from openpyxl import load_workbook
from . models import *

# Create your views here.

# Create a new Requirement for url '/requirements'
## In this, client makes a reqruirement for the design.

@login_required
def requirements(request):
    return render(request, "requirements.html")


# Webscrapping
# Update Scrapping Material Datas in material table
## In Excel, last column is state of material. 0: deleted, 1: existed, 2: created
## In Table, If state is True, material is used. If False, deleted.

def scrap_history(request):
    categories = Category.objects.all()

    for category in categories:
        filename = finders.find('upload/scrap_data/' + category.name + '.xlsx')

        wb = load_workbook(filename)

        sheets = category.sheetname.split(',')
        
        for sh in sheets:
            sheet = wb[sh]
            i = 2
            for row in sheet.iter_rows(min_row=2, min_col=2):
                if row[10].value == 2:
                    material = Material(
                        title = row[0].value,
                        image = row[1].value,
                        specification = row[2].value,
                        ref = row[3].value,
                        price = row[4].value,
                        description = row[5].value,
                        advantage = row[6].value,
                        documentation = row[7].value,
                        features = row[8].value,
                        url = row[9].value,
                        category_id = category.id,
                        state = 1
                    )

                    material.save()

                    sheet.cell(row=i, column=12).value = 1
                
                i = i + 1
            
            wb.save(filename)
            
    return render(request, 'index.html')